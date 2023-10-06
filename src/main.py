import os 
import sys
import time
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.chart import BarChart, Reference, Series, BarChart3D,PieChart, PieChart3D, LineChart,RadarChart,ScatterChart,AreaChart,StockChart, BubbleChart, SurfaceChart,DoughnutChart
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.chart.label import DataLabelList

import locale
locale.setlocale(locale.LC_ALL, 'ca_ES.utf8') #Catalan
from tkinter import Tk, Button, messagebox, Label
from tkcalendar import Calendar, DateEntry
import datetime

#no importar los warnings
import warnings
warnings.filterwarnings("ignore")

now = datetime.datetime.now()



# Define una función para convertir la fecha si cumple con el formato
def convertir_fecha(fecha):
    try:
        return pd.to_datetime(fecha, format='%d/%m/%Y')
    except ValueError:
        return fecha


########### Comentarios ###########
# 1. En el excel de plantilla, en la hoja de tecnologias, hay una tabla con las columnas [Fecha,Tecnologies,Producció OK,Producció KO,Total Producció]
###################################


def read_excel(path, sheet_name):
    try:
        df = pd.read_excel(path, sheet_name=sheet_name)
    except FileNotFoundError:
        raise FileNotFoundError(f"L'arxiu '{path}' no existeix. Verifiqueu la ruta i el nom de l'arxiu.")
    except Exception as e:
        raise Exception(f"S'ha produït un error en llegir l'arxiu '{path}': {str(e)}")

    try:
        # Eliminar files amb tots els valors NaN
        df = df.dropna(how='all')
        
        # Sustituir els valors NaN per 0
        df = df.fillna(0)
    except Exception as e:
        raise Exception(f"S'ha produït un error en netejar les dades: {str(e)}")

    return df


def generate_desplegament_total(df):
    try:
        # Hacer una copia del dataframe para no afectar el original
        df = df.copy()

        df['Fecha'] = df['Fecha'].apply(convertir_fecha)
        
        # En la columna de fecha, se debe poner el formato de año y mes
        df['Fecha'] = pd.to_datetime(df['Fecha']).dt.strftime('%Y-%m-%B')

        # Eliminar la columna de Urgentes
        df = df.drop(columns=['Urgents'])
        
        # Agrupar por fecha y sumar las columnas de Producció OK y Producció KO
        df = df.groupby(['Fecha']).sum()
        
        # Resetear el índice para que la columna de fecha sea una columna normal
        df = df.reset_index()
        
        # Ordenar por años y mes en orden de meses
        df = df.sort_values(by=['Fecha'], ascending=True)
        
        # Eliminar de la fecha el año y mes para que solo quede el nombre del mes
        df['Fecha'] = df['Fecha'].str.split('-').str[2]
        
        # Poner en mayúsculas la primera letra del mes
        df['Fecha'] = df['Fecha'].str.capitalize()
        
        # Por si acaso, sumar en la columna de Total Producció los valores de Producció OK y Producció KO
        df['Total Producció'] = df['Producció OK'] + df['Producció KO']
        
        # Poner la abreviación de los meses y las letras en minúsculas
        df['Fecha'] = df['Fecha'].str.slice(stop=3).str.lower()
        
    except Exception as e:
        raise Exception(f"S'ha produït un error en generar el desplegament total: {str(e)}")

    return df


def generate_desplegamnet_mes(df, mes):
    try:
        df = df.copy()

        df['Fecha'] = df['Fecha'].apply(convertir_fecha)

        # En la columna de fecha, se debe poner el formato de año y mes
        df['Fecha'] = pd.to_datetime(df['Fecha']).dt.strftime('%Y-%m')
        
        # Eliminar la columna de Urgentes
        df = df.drop(columns=['Urgents'])
        
        # Filtrar por mes
        df = df[df['Fecha'] == mes]
        
        # Agrupar por tecnologia y sumar las columnas de Producció OK y Producció KO
        df = df.groupby(['Tecnologies']).sum()
        
        # Resetear el índice para que la columna de tecnologia sea una columna normal
        df = df.reset_index()
        
        # Ordenar el dataframe de mayor a menor por la columna de Total Producció
        df = df.sort_values(by=['Total Producció'], ascending=False)
        
        # Eliminar del dataframe las filas que tengan el valor de Total Producció igual a 0
        df = df[df['Total Producció'] != 0]
        
        # Por si acaso, sumar en la columna de Total Producció los valores de Producció OK y Producció KO
        df['Total Producció'] = df['Producció OK'] + df['Producció KO']
        
    except Exception as e:
        raise Exception(f"S'ha produït un error en generar el desplegament per al mes {mes}: {str(e)}")
    
    return df

def generate_total_tecnologia(df):
    try:
        df = df.copy()
        # Agrupar por tecnologias y sumar el total es decir creamos un nuevo dataframe con las columnas [Tecnologies,Total Producción]
        df = df.groupby(['Tecnologies']).sum()

        # Resetear el index para que la columna de fecha sea una columna normal
        df = df.reset_index()

        # Ordenar el dataframe de mayor a menor por la columna de Total Producció
        df = df.sort_values(by=['Total Producció'],ascending=True)

        #Eliminar del dataframe las filas que tengan el valor de Total Producció igual a 0
        df = df[df['Total Producció'] != 0]

        df  = df[['Tecnologies','Total Producció']]
    except Exception as e:
        raise Exception(f"S'ha produït un error en generar el total de tecnologies: {str(e)}")
    return df

def generate_total_tecnologia_mes(df,mes):
    try:
        df = df.copy()

        df['Fecha'] = df['Fecha'].apply(convertir_fecha)
        #En la columa de fecha, se debe de poner el formato de año y mes
        df['Fecha'] = pd.to_datetime(df['Fecha']).dt.strftime('%Y-%m')

        #Seleccionar datos del mes
        df = df[df['Fecha'] == mes]

        # Agrupar por tecnologias y sumar el total es decir creamos un nuevo dataframe con las columnas [Tecnologies,Total Producción]
        df = df.groupby(['Tecnologies']).sum()

        # Resetear el index para que la columna de fecha sea una columna normal
        df = df.reset_index()

        # Ordenar el dataframe de mayor a menor por la columna de Total Producció
        df = df.sort_values(by=['Total Producció'],ascending=True)

        #Eliminar del dataframe las filas que tengan el valor de Total Producció igual a 0
        df = df[df['Total Producció'] != 0]

        df  = df[['Tecnologies','Total Producció']]
    except Exception as e:
            raise Exception(f"S'ha produït un error en generar el total de tecnologies per mes: {str(e)}")
   
    return df


def generate_total_desplegament_tecnologia_mes(df):
    try:
        df = df.copy()

        df['Fecha'] = df['Fecha'].apply(convertir_fecha)

        #En la columa de fecha, se debe de poner el formato de año y mes
        df['Fecha'] = pd.to_datetime(df['Fecha']).dt.strftime('%Y-%m-%B')

         #Crearemos un nuevo dataset con las columnas fechas y una nueva columna por cada tecnologia del dataset original
        df = df.pivot_table(index=['Fecha'],columns=['Tecnologies'],values=['Total Producció'],aggfunc=np.sum)

    
        #Resetear el index para que la columna de fecha sea una columna normal
        df = df.reset_index()

        #ponemos todos los NaN a 0
        df = df.fillna(0)

        #ordenar por años y mes en orden de meses
        df = df.sort_values(by=['Fecha'],ascending=True)

        #eliminar de la fecha el año y mes para que solo quede el nombre del mes
        df['Fecha'] = df['Fecha'].str.split('-').str[2]

        #Abreviar los nombres de los meses
        df['Fecha'] = df['Fecha'].str.slice(stop=3)


        #eliminar la cabecera de las columnas
        df.columns = df.columns.droplevel(0)

        #creamos una columna nueva con el total general de produccion
        df['Total General'] = df.sum(axis=1)
    except Exception as e:
            raise Exception(f"S'ha produït un error en generar el total de desplegaments per tecnologia per mes: {str(e)}")

    


    return df

def generate_total_desplegament_DevOps_mes(df):
    try:
        df = df.copy()
        #Tomar solo los valores donde tecnologia sea igual a Devops
        df = df[df['Tecnologies'] == 'Devops']

        #elimina la columna "Producció OK"
        df = df.drop(['Producció OK'],axis=1)

        #eliminar la columna de Urgentes
        df = df.drop(['Urgents'],axis=1)

        df['Fecha'] = df['Fecha'].apply(convertir_fecha)

         #En la columa de fecha, se debe de poner el formato de año y mes
        df['Fecha'] = pd.to_datetime(df['Fecha']).dt.strftime('%Y-%m-%B')

        #Agrupar por fecha y sumar las columnas de Producció OK y Producció KO
        df = df.groupby(['Fecha']).sum()

        #Resetear el index para que la columna de fecha sea una columna normal
        df = df.reset_index()

        #ordenar por años y mes en orden de meses
        df = df.sort_values(by=['Fecha'],ascending=True)

        #eliminar de la fecha el año y mes para que solo quede el nombre del mes
        df['Fecha'] = df['Fecha'].str.split('-').str[2]

        # poner solo las 3 primeras letras
        df['Fecha'] = df['Fecha'].str.slice(stop=3)

        #Agregar una columna que sea % de KO sobre total de produccion el calculo es Producció KO / Total Producció, mostrar sin decimales
        df['% KO'] = (df['Producció KO'] / df['Total Producció']) * 100

        #eliminar los decimales de la columna % KO
        df['% KO'] = df['% KO'].astype(int)

        # las columnas son Fecha, Producció KO, Total Producció, % KO, quiero cambiar la posicion de Producció KO, Total Producció, para que quede Fecha, Total Producció, Producció KO, % KO
        df = df[['Fecha','Total Producció','Producció KO','% KO']]
    except Exception as e:
            raise Exception(f"S'ha produït un error en generar el total DevOps  per mes: {str(e)}")

    return df


def generate_total_desplegament_Urgente_mes(df):
    try:
        df = df.copy()

        #eliminar la columna "Producció OK"
        df = df.drop(['Producció OK'],axis=1)

        #eliminar la columna "Producció KO"
        df = df.drop(['Producció KO'],axis=1)

        df['Fecha'] = df['Fecha'].apply(convertir_fecha)

        #En la columa de fecha, se debe de poner el formato de año y mes
        df['Fecha'] = pd.to_datetime(df['Fecha']).dt.strftime('%Y-%m-%B')

        #Agrupar por fecha y sumar las columnas de Producció OK y Producció KO
        df = df.groupby(['Fecha']).sum()

        #Resetear el index para que la columna de fecha sea una columna normal
        df = df.reset_index()

        #ordenar por años y mes en orden de meses
        df = df.sort_values(by=['Fecha'],ascending=True)

        #eliminar de la fecha el año y mes para que solo quede el nombre del mes
        df['Fecha'] = df['Fecha'].str.split('-').str[2]

        # abreviar los meses dejar solo las 3 primeras letras
        df['Fecha'] = df['Fecha'].str.slice(stop=3)


        #Agregar una columna que sea % de Urgentes sobre total de produccion el calculo es Urgentes / Total Producció, mostrar sin decimales
        df['% Urgents'] = (df['Urgents'] / df['Total Producció']) * 100

        #eliminar los decimales de la columna % Urgentes
        df['% Urgents'] = df['% Urgents'].astype(int)

    except Exception as e:
            raise Exception(f"S'ha produït un error en generar el total Urgentes  per mes: {str(e)}")

    return df

 
def generate_total_per_mes(df, year):
    try:
        # Copiar el DataFrame para no modificar el original
        df_copy = df.copy()

        # Filtrar las filas donde 'Tecnologies' no sea igual a 'Total'
        df_copy = df_copy[df_copy['Tecnologies'] != 'Total']

        df_copy['Fecha'] = df_copy['Fecha'].apply(convertir_fecha)

        # Convertir la columna 'Fecha' al tipo datetime
        df_copy['Fecha'] = pd.to_datetime(df_copy['Fecha'])

        # Filtrar por el año deseado
        df_copy = df_copy[df_copy['Fecha'].dt.year == year]

        # Función para obtener el nombre del mes en formato deseado
        def obtener_nombre_mes(date):
            meses = ['gen', 'febr', 'març', 'abr', 'maig', 'juny', 'juliol', 'ag', 'set', 'oct', 'nov', 'des']
            return meses[date.month - 1]

        # Aplicar la función para obtener el nombre del mes
        df_copy['Mes'] = df_copy['Fecha'].apply(obtener_nombre_mes)

        # Limpiar y convertir las columnas 'Producció OK' y 'Producció KO' a valores numéricos
        df_copy['Producció OK'] = pd.to_numeric(df_copy['Producció OK'], errors='coerce')
        df_copy['Producció KO'] = pd.to_numeric(df_copy['Producció KO'], errors='coerce')

        # Seleccionar solo las columnas 'Mes', 'Producció OK' y 'Producció KO'
        df_copy = df_copy[['Mes', 'Producció OK', 'Producció KO']]

        # Agrupar por mes y sumar los valores
        df_copy = df_copy.groupby('Mes').sum().reset_index()

        # Crear una columna 'Numero Mes' para ordenar correctamente
        meses_ordenados =['gen', 'febr', 'març', 'abr', 'maig', 'juny', 'juliol', 'ag', 'set', 'oct', 'nov', 'des']
        df_copy['Numero Mes'] = df_copy['Mes'].apply(lambda x: meses_ordenados.index(x) + 1)

        # Ordenar por el número de mes
        df_copy = df_copy.sort_values(by='Numero Mes')

        # Eliminar la columna 'Numero Mes'
        df_copy = df_copy.drop(columns='Numero Mes')

        #quitar el decimal a las columnas de Producció OK y Producció KO
        df_copy['Producció OK'] = df_copy['Producció OK'].astype(int)
        df_copy['Producció KO'] = df_copy['Producció KO'].astype(int)
    except Exception as e:
            raise Exception(f"S'ha produït un error en generar el total per mes comparacio anys: {str(e)}")
 
    return df_copy

 

def generate_comparacio_anys(df_anterior,df_actual,n=0,m=1):
    try:
        df_anterior = df_anterior.rename(columns={'Producció OK':'Producció OK ' + str(n),'Producció KO':'Producció KO ' + str(n)})

        df_actual = df_actual.rename(columns={'Producció OK':'Producció OK ' + str(m),'Producció KO':'Producció KO ' + str(m)})

        #ahora vamos a unir los dos dataframes
        df = pd.merge(df_anterior,df_actual,how='inner',on='Mes')

        #agregar una colunma que sea el total de produccion de los dos años
        df['Total Producció ' + str(n)] = df['Producció OK ' + str(n)] + df['Producció KO ' + str(n)]
        df['Total Producció ' + str(m)] = df['Producció OK ' + str(m)] + df['Producció KO ' + str(m)]

        #eliminar las columnas de Producció OK y Producció KO del ano anterior
        df = df.drop(columns=['Producció OK ' + str(m),'Producció KO ' + str(m)])
    except Exception as e:
            raise Exception(f"S'ha produït un error en generar el total per mes comparacio anys durant la agrupació: {str(e)}")

    return df




def excel_style(ws):
    # Define the font and alignment for the header
    font = Font(name='Arial', size=12, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='ffffff')
    alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)

    # Define the fill for the header
    fill = PatternFill(fill_type='solid', start_color='000000', end_color='000000')

    # Apply the style to the header
    for cell in ws["1:1"]:
        ws.column_dimensions[cell.column_letter].width = 20
        cell.font = font
        cell.alignment = alignment
        cell.fill = fill


    #Colores a las filas intercaladas
    # Define the fill color for odd rows
    fill = PatternFill(start_color='b3d2ff', end_color='FFC000', fill_type='solid')

    # Apply the style to odd rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = fill
    
    return ws
    


def generate_graphic(df, doc_excel, hoja=None, title=None, chart_type=None):
    # Si el archivo Excel existe, se abre y se agrega la hoja de trabajo; si no existe, se crea el archivo Excel
    if os.path.isfile(doc_excel):
        wb = openpyxl.load_workbook(doc_excel)
        if hoja in wb.sheetnames:
            ws = wb[hoja]
        else:
            ws = wb.create_sheet(hoja)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = hoja

    # Borrar la tabla de la hoja de trabajo y agregar el DataFrame
    ws.delete_rows(1, ws.max_row)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    ws = excel_style(ws)

    # Nombre de la hoja
    ws.title = hoja

    # Guardar el archivo Excel
    wb.save(doc_excel)

    # Generar gráfico si se especifica el tipo de gráfico
    if chart_type == 'circular':
        generate_circular_graphic(df, wb, hoja, title)
    elif chart_type == 'horizontal':
        generate_horizontal_graphic(df, wb, hoja, title)
    elif chart_type == 'barras_lineal':
        generate_graphic_barras_lineal(df, wb, hoja, title)

def generate_circular_graphic(df, wb, hoja=None, title=None):
    # Agregar aquí la lógica para generar un gráfico circular en la hoja especificada
    pass

def generate_horizontal_graphic(df, wb, hoja=None, title=None):
    # Agregar aquí la lógica para generar un gráfico horizontal en la hoja especificada
    pass

def generate_graphic_barras_lineal(df, wb, hoja=None, title=None):
    # Agregar aquí la lógica para generar un gráfico de barras o lineal en la hoja especificada
    pass





'''def generate_graphic(df,doc_excel,hoja=None,title=None,x_axis=None):
    if os.path.isfile(doc_excel):
        wb = openpyxl.load_workbook(doc_excel)
        #compobar si la hoja existe
        if hoja in wb.sheetnames:
            ws = wb[hoja]
        else:   
                ws = wb.create_sheet(hoja)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = hoja

    ws.delete_rows(1,ws.max_row)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    ws = excel_style(ws)
   
    #nombre de la hoja
    ws.title = hoja
    
    # Guardar el archivo
    wb.save(doc_excel)


def generate_circular_graphic(df,doc_excel,hoja=None,title=None):

    #si el archivo excel existe, se abre y agrega la hoja de trabajo, si no existe, se crea el archivo excel
    if os.path.isfile(doc_excel):
        wb = openpyxl.load_workbook(doc_excel)
        if hoja in wb.sheetnames:
            ws = wb[hoja]
        else:
            ws = wb.create_sheet(hoja)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = hoja

    # borra la tabla de la hoja de trabajo y agregar el dataframe
    ws.delete_rows(1,ws.max_row)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    ws = excel_style(ws)
    
    #guardar el archivo
    wb.save(doc_excel)


def generate_horizontal_graphic(df,doc_excel,hoja=None,title=None):

    #si el archivo excel existe, se abre y agrega la hoja de trabajo, si no existe, se crea el archivo excel
    if os.path.isfile(doc_excel):
        wb = openpyxl.load_workbook(doc_excel)
        ws = wb[hoja]
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = hoja

    # borrar la tabla de la hoja de trabajo y agregar el dataframe
    ws.delete_rows(1,ws.max_row)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    ws = excel_style(ws)
   
    #nombre de la hoja
    ws.title = hoja

    # Guardar el archivo
    wb.save(doc_excel)

def generate_graphic_barras_lineal(df,doc_excel,hoja=None,title=None):
    #si el archivo excel existe, se abre y agrega la hoja de trabajo, si no existe, se crea el archivo excel
    if os.path.isfile(doc_excel):
        wb = openpyxl.load_workbook(doc_excel)
        #Ubicarnos en la hoja de trabajo , PERO NO crearla
        ws = wb[hoja]
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = hoja

    # borrar la tabla de la hoja de trabajo y agregar el dataframe
    ws.delete_rows(1,ws.max_row)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    ws = excel_style(ws)

    #nombre de la hoja
    ws.title = hoja

    # Guardar el archivo
    wb.save(doc_excel)'''


#Generar una funcion para mostar un calendario por pantalla que solo permita seleccionar un mes y un año

def get_date():
    """Muestra un calendario para que el usuario seleccione una fecha."""
    root = Tk()
    root.geometry("300x350")
   
    root.title("Seleccionar Fecha")
   
    def get_selected_date():
        """Obtiene la fecha seleccionada por el usuario y cierra la ventana."""
        nonlocal selected_date
        selected_date = cal.selection_get()
        root.destroy()

    def cancel():
        """Cierra la ventana sin seleccionar una fecha."""
        nonlocal selected_date
        selected_date = None
        root.destroy()
        exit()

    #Poner un titulo en la ventana
    title_label = Label(root, text="Seleccionar Data", font=("Helvetica", 16), fg="white", bg="black")
    title_label.pack(pady=10)
    
    selected_date = None
    
    cal = Calendar(root, selectmode="day", year=now.year , month=now.month, day=now.day,maxdate=now)
    cal.pack(pady=20)


    btn_ok = Button(root, text="OK", command=get_selected_date)
    btn_ok.pack(side="left", pady=10, padx=10)

    btn_cancel = Button(root, text="CANCEL·LAR", command=cancel)
    btn_cancel.pack(side="right", pady=10, padx=10)

    root.protocol("WM_DELETE_WINDOW", lambda: messagebox.showerror("Error", "Heu de seleccionar una data"))

    root.mainloop()

    if selected_date is None:
        raise ValueError("No s'ha seleccionat cap data")
    return selected_date
    
    

def main():
    path = "Plantilla.xlsx"
    sheet_name = 'Tecnologies'

    #Mostrar por pantalla
    print("#########################################")
    print("######### Generacio de grafics ###########")
    print("#########################################")
    print()
    print("Per defecte es mostraran els grafics de l'ultim mes")
    print("Si vols seleccionar un altre mes, premi la tecla 'N' i introdueix el mes i l'any")
    print()
    print("Si vols CONTINUAR, premi la tecla 'S'")


    #Preguntar al usuario si quiere seleccionar un mes
    seleccionar_mes = input("Vols seleccionar un mes? (N): ")

    if seleccionar_mes == 'N' or seleccionar_mes == 's':
        #Obtener la fecha seleccionada por el usuario mostrar lista de meses 
        print("#########################################")
        print("01. Gener")
        print("02. Febrer")
        print("03. Març")
        print("04. Abril")
        print("05. Maig")
        print("06. Juny")
        print("07. Juliol")
        print("08. Agost")
        print("09. Setembre")
        print("10. Octubre")
        print("11. Novembre")
        print("12. Desembre")
        print("#########################################")
        mes = int(input("Selecciona el mes: "))
        year = int(input("Selecciona l'any: "))

        #incluimos el 0 en el mes para que quede 01,02,03,04,05,06,07,08,09
        if mes < 10:
            fecha = str(year) + '-' + '0' + str(mes)
        else:
            fecha = str(year) + '-' + str(mes)

    else:
        #Obtener la fecha actual
        fecha = str(now.year) + '-' + str(now.month)

    print("#########################################")
    print("Generant grafics...")
    print("Fecha seleccionada: " + fecha)
    print("#########################################")

    df = read_excel(path,sheet_name)

    generate_graphic(generate_total_desplegament_tecnologia_mes(df),'grafico.xlsx','Total','Evolució desplegaments per tecnologia')
    
    generate_graphic(generate_desplegament_total(df),'grafico.xlsx','Total desplegaments',"Desplegaments - Evolució mensual","Mesos")
    generate_graphic(generate_desplegamnet_mes(df,fecha),'grafico.xlsx','Total desplegaments mes',"Deplegaments -" + fecha,"Tecnologies")

    generate_graphic(generate_total_tecnologia(df),'grafico.xlsx','Total tecnologia',"Evolució Mensual per volum")
    generate_graphic(generate_total_tecnologia_mes(df,fecha),'grafico.xlsx','Total tecnologia mes',"Evolució Mensual per volum -" + fecha)
    
    generate_graphic( generate_total_desplegament_Urgente_mes(df),'grafico.xlsx','% KO Urgentes',"% Peticions Urgents")
    generate_graphic(generate_total_desplegament_DevOps_mes(df), 'grafico.xlsx', '% KO DevOps', "% Peticions DevOps")

    
    
    sheet_name = 'Master'

    df_anterior = generate_total_per_mes(read_excel(path,sheet_name),now.year-1)
    df_actual = generate_total_per_mes(df,now.year)

    generate_graphic(generate_comparacio_anys(df_actual,df_anterior,now.year,now.year-1),'grafico.xlsx','Comparació anys',"Comparació anys","Mesos")

    print("#########################################")
    print("Grafiques generades correctament")
    print("Reviseu el fitxer Excel")
    print("#########################################")

if __name__ == '__main__':
    try:
        main()
        #press enter to exit
        input("Premeu una tecla per a continuar...")
    except Exception as e:
        messagebox.showerror("Error", str(e))
        sys.exit()
